from flask import Flask, render_template, request, redirect, url_for
import os
import openpyxl
from werkzeug.utils import secure_filename

app = Flask(__name__)

# Путь к загруженным Excel файлам
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@app.route('/a')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return redirect(request.url)

    file = request.files['file']

    if file.filename == '':
        return redirect(request.url)

    if file:
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)

        # Открываем загруженный Excel файл и парсим данные
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        data = []

        for univer in sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=12):
            univer_data = []
            for cell in univer:
                # Преобразуем значение ячейки в строку и убираем лишние символы
                cell_value = str(cell.value).strip()
                if cell_value != 'None':
                    univer_data.append(cell_value)
            data.append(univer_data)

        # Институт\Факультет Проходимся по строкам и столбцам, которые вам нужны, и добавляем их в список
        for Institut in sheet.iter_rows(min_row=2, max_row=2, min_col=1, max_col=23):
            Institut_data = []
            for cell in Institut:
                # Преобразуем значение ячейки в строку и убираем лишние символы
                cell_value = str(cell.value).strip()
                if cell_value != 'None':
                    Institut_data.append(cell_value)
            data.append(Institut_data)

        # Форма обучения Проходимся по строкам и столбцам, которые вам нужны, и добавляем их в список
        for FormOfEdc in sheet.iter_rows(min_row=3, max_row=3, min_col=1, max_col=16):
            FormOfEdc_data = []
            for cell in FormOfEdc:
                # Преобразуем значение ячейки в строку и убираем лишние символы
                cell_value = str(cell.value).strip()
                if cell_value != 'None':
                    FormOfEdc_data.append(cell_value)
            data.append(FormOfEdc_data)

        #  Уровень образования Проходимся по строкам и столбцам, которые вам нужны, и добавляем их в список
        for Uroven in sheet.iter_rows(min_row=4, max_row=4, min_col=1, max_col=16):
            Uroven_data = []
            for cell in Uroven:
                # Преобразуем значение ячейки в строку и убираем лишние символы
                cell_value = str(cell.value).strip()
                if cell_value != 'None':
                    Uroven_data.append(cell_value)
            data.append(Uroven_data)

        # Расписание заняти и дата Проходимся по строкам и столбцам, которые вам нужны, и добавляем их в список
        for row in sheet.iter_rows(min_row=6, max_row=6, min_col=1, max_col=44):
            row_data = []
            for cell in row:
                # Преобразуем значение ячейки в строку и убираем лишние символы
                cell_value = str(cell.value).strip()
                if cell_value != 'None':
                    row_data.append(cell_value)
            data.append(row_data)

        #  Код направления Проходимся по строкам и столбцам, которые вам нужны, и добавляем их в список
        for kod in sheet.iter_rows(min_row=7, max_row=7, min_col=1, max_col=44):
            kod_data = []
            for cell in kod:
                # Преобразуем значение ячейки в строку и убираем лишние символы
                cell_value = str(cell.value).strip()
                if cell_value != 'None':
                    kod_data.append(cell_value)
            data.append(kod_data)

        #  Группа направления Проходимся по строкам и столбцам, которые вам нужны, и добавляем их в список
        for Group in sheet.iter_rows(min_row=8, max_row=8, min_col=1, max_col=44):
            Group_data = []
            for cell in Group:
                # Преобразуем значение ячейки в строку и убираем лишние символы
                cell_value = str(cell.value).strip()
                if cell_value != 'None':
                    Group_data.append(cell_value)
            data.append(Group_data)

        return render_template('result.html', data=data)

if __name__ == '__main__':
    app.run(debug=True)
