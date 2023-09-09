import openpyxl
from flask import Flask, request, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from marshmallow import Schema, fields
from flask_admin import Admin
from flask_admin.contrib.sqla import ModelView

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://your_username:your_password@localhost/your_database'
db = SQLAlchemy(app)
migrate = Migrate(app, db)
admin = Admin(app, name='Admin Panel', template_mode='bootstrap3')

# Определение моделей
class Teacher(db.Model):
    __tablename__ = 'new_teachers'

    id = db.Column(db.Integer, primary_key=True)
    surname = db.Column(db.String(255), nullable=False)
    name = db.Column(db.String(255), nullable=False)
    patronymic = db.Column(db.String(255))

class Item(db.Model):
    __tablename__ = 'new_items'

    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(255), nullable=False)

class LessonType(db.Model):
    __tablename__ = 'new_lesson_type'

    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(255), nullable=False)

class Frame(db.Model):
    __tablename__ = 'new_frames'

    id = db.Column(db.Integer, primary_key=True)
    address = db.Column(db.String(255), nullable=False)

class Audience(db.Model):
    __tablename__ = 'new_audiences'

    id = db.Column(db.Integer, primary_key=True)
    floor_number = db.Column(db.Integer, nullable=False)
    room_number = db.Column(db.String(10), nullable=False)
    building_id = db.Column(db.Integer, db.ForeignKey('new_frames.id', ondelete='SET NULL'))
    building = db.relationship('Frame', backref='audiences')

class Time(db.Model):
    __tablename__ = 'new_time'

    id = db.Column(db.Integer, primary_key=True)
    start_time = db.Column(db.Time)
    end_time = db.Column(db.Time)

class Group(db.Model):
    __tablename__ = 'new_groups'

    id = db.Column(db.Integer, primary_key=True)
    faculty = db.Column(db.String(255), nullable=False)
    well = db.Column(db.Integer, nullable=False)
    number = db.Column(db.Integer, nullable=False)

class Schedule(db.Model):
    __tablename__ = 'new_schedule'

    id = db.Column(db.Integer, primary_key=True)
    data_les = db.Column(db.Date)
    week = db.Column(db.String(255))
    id_group = db.Column(db.Integer, db.ForeignKey('new_groups.id', ondelete='SET NULL'))
    group = db.relationship('Group', backref='schedules')
    id_lesson_type = db.Column(db.Integer, db.ForeignKey('new_lesson_type.id', ondelete='SET NULL'))
    lesson_type = db.relationship('LessonType', backref='schedules')
    id_item = db.Column(db.Integer, db.ForeignKey('new_items.id', ondelete='SET NULL'))
    item = db.relationship('Item', backref='schedules')
    id_teacher = db.Column(db.Integer, db.ForeignKey('new_teachers.id', ondelete='SET NULL'))
    teacher = db.relationship('Teacher', backref='schedules')
    id_audience = db.Column(db.Integer, db.ForeignKey('new_audiences.id', ondelete='SET NULL'))
    audience = db.relationship('Audience', backref='schedules')

# Определение схем для маршализации данных
class TeacherSchema(Schema):
    id = fields.Integer()
    surname = fields.String()
    name = fields.String()
    patronymic = fields.String()

class ItemSchema(Schema):
    id = fields.Integer()
    name = fields.String()

class LessonTypeSchema(Schema):
    id = fields.Integer()
    name = fields.String()

#
class TeacherView(ModelView):
    column_display_pk = True  # Отображать первичные ключи
    column_searchable_list = ['surname', 'name']  # Поля для поиска

# Добавляем представления (views) для моделей
admin.add_view(ModelView(Teacher, db.session, name='Teachers'))
admin.add_view(ModelView(Item, db.session, name='Items'))
admin.add_view(ModelView(LessonType, db.session, name='Lesson Types'))
admin.add_view(ModelView(Frame, db.session, name='Frames'))
admin.add_view(ModelView(Audience, db.session, name='Audiences'))
admin.add_view(ModelView(Time, db.session, name='Times'))
admin.add_view(ModelView(Group, db.session, name='Groups'))
admin.add_view(ModelView(Schedule, db.session, name='Schedules'))

# Эндпоинт для загрузки Excel-файла с расписанием
@app.route('/upload-schedule', methods=['POST'])
def upload_schedule():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400

    if file:
        try:
            # Открываем Excel-файл с помощью библиотеки openpyxl
            wb = openpyxl.load_workbook(file)
            sheet = wb.active

            # Проходим по строкам Excel-файла и извлекаем данные
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                # Здесь вы можете извлекать и сохранять данные в вашу базу данных
                # Пример:
                group = row[5].value
                # day = row[6].value
                # time = row[7].value
                # discipline = row[8].value
                # location = row[9].value
                # ...

            # Сохраняем изменения в базе данных
            db.session.commit()

            return jsonify({'message': 'Schedule uploaded successfully'}), 200
        except Exception as e:
            return jsonify({'error': str(e)}), 500

# Добавление других эндпоинтов для получения данных
@app.route('/teachers', methods=['GET'])
def get_teachers():
    teachers = Teacher.query.all()
    teacher_schema = TeacherSchema(many=True)
    result = teacher_schema.dump(teachers)
    return jsonify(result)

@app.route('/items', methods=['GET'])
def get_items():
    items = Item.query.all()
    item_schema = ItemSchema(many=True)
    result = item_schema.dump(items)
    return jsonify(result)

@app.route('/lesson-types', methods=['GET'])
def get_lesson_types():
    lesson_types = LessonType.query.all()
    lesson_type_schema = LessonTypeSchema(many=True)
    result = lesson_type_schema.dump(lesson_types)
    return jsonify(result)

# Добавление других эндпоинтов для остальных моделей аналогичным образом

if __name__ == '__main__':
    app.run(debug=True)
